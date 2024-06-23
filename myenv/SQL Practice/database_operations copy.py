import os
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import sqlite3
import openpyxl
from openpyxl import Workbook, load_workbook


# Paths
script_dir = os.path.dirname(__file__)
db_path = os.path.join(script_dir, 'Books_database.db')
sql_file_path = os.path.join(script_dir, 'database_setup.sql')
excel_file_path = os.path.join(script_dir, 'books_database.xlsx')

# Function to execute SQL script
def execute_sql_script(sql_file):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    if not os.path.exists(sql_file):
        raise FileNotFoundError(f"SQL file '{sql_file}' not found.")

    with open(sql_file, 'r') as f:
        sql_script = f.read()
        cursor.executescript(sql_script)
        conn.commit()

    conn.close()

# Function to retrieve authors from database
def get_authors():
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT author_name FROM Authors ORDER BY author_name")
    authors = [row[0] for row in cursor.fetchall()]
    conn.close()
    return authors

# Function to add a new author to the database
def add_new_author(first_name, last_name):
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Insert new author
        author_name = f"{first_name} {last_name}"
        cursor.execute("INSERT INTO Authors (author_name) VALUES (?)", (author_name,))
        conn.commit()
        conn.close()
        
        messagebox.showinfo("Success", f"Author '{author_name}' added successfully.")
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Error adding author: {e}")

# Function to remove an author from the database
def remove_author(author_name):
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Check if author exists
        cursor.execute("SELECT author_id FROM Authors WHERE author_name=?", (author_name,))
        author_id = cursor.fetchone()
        
        if author_id:
            author_id = author_id[0]
            # Delete author and associated books (if any)
            cursor.execute("DELETE FROM Authors WHERE author_id=?", (author_id,))
            conn.commit()
            messagebox.showinfo("Success", f"Author '{author_name}' removed successfully.")
        else:
            messagebox.showwarning("Author not found", f"Author '{author_name}' not found.")
        
        conn.close()
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Error removing author: {e}")

# Function to add a book to the database
def add_book(title, author):
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Check if author exists
        cursor.execute("SELECT author_id FROM Authors WHERE author_name=?", (author,))
        author_id = cursor.fetchone()
        
        if author_id:
            author_id = author_id[0]
        else:
            # Insert new author
            cursor.execute("INSERT INTO Authors (author_name) VALUES (?)", (author,))
            author_id = cursor.lastrowid
        
        # Insert book
        cursor.execute("INSERT INTO Books (title, author_id) VALUES (?, ?)", (title, author_id))
        
        conn.commit()
        conn.close()
        
        messagebox.showinfo("Success", "Book added successfully.")
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Error adding book: {e}")

# Function to remove a book from the database
def remove_book(title):
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Books WHERE title=?", (title,))
        book = cursor.fetchone()

        if book:
            cursor.execute("DELETE FROM Books WHERE title=?", (title,))
            conn.commit()
            messagebox.showinfo("Success", "Book removed successfully.")
        else:
            messagebox.showwarning("Book not found", f"Book with title '{title}' not found.")
    except sqlite3.Error as e:
        messagebox.showerror("Database Error", f"Error removing book: {e}")
    finally:
        conn.close()

# Function to handle adding a book when Add Book button is clicked
def on_add_book():
    title = entry_title.get().strip()
    author = combo_author.get().strip()
    
    if not title:
        messagebox.showwarning("Missing Information", "Please enter a title.")
        return
    
    if not author:
        messagebox.showwarning("Missing Information", "Please enter an author.")
        return
    
    add_book(title, author)
    # Refresh authors list after adding new author
    combo_author['values'] = get_authors()

# Function to handle removing an author when Remove Author button is clicked
def on_remove_author():
    author_name = combo_author.get().strip()
    
    if not author_name:
        messagebox.showwarning("Missing Information", "Please select an author to remove.")
        return
    
    remove_author(author_name)
    # Refresh authors list after removing author
    combo_author['values'] = get_authors()

# Function to create a new window for adding a new author
def add_new_author_window():
    new_author_window = tk.Toplevel()
    new_author_window.title("Add New Author")
    
    frame = ttk.Frame(new_author_window, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)
    
    ttk.Label(frame, text="Author Name:").grid(row=0, column=0, padx=10, pady=5)
    author_name_entry = ttk.Entry(frame, width=30)
    author_name_entry.grid(row=0, column=1, padx=10, pady=5)
    
    # Function to handle adding new author
    def submit_author():
        author_name = author_name_entry.get().strip()
        
        if not author_name:
            messagebox.showwarning("Missing Information", "Please enter the author's name.")
            return
        
        add_new_author(author_name)
        new_author_window.destroy()
    
    ttk.Button(frame, text="Add Author", command=submit_author).grid(row=1, column=0, columnspan=2, pady=10)

# Function to handle removing a book when Remove Book button is clicked
def on_remove_book():
    title = entry_title.get().strip()
    
    if not title:
        messagebox.showwarning("Missing Information", "Please enter a title to remove.")
        return
    
    remove_book(title)

# Function to create a new window for adding a new author
def add_new_author_window():

    new_author_window = tk.Toplevel()
    new_author_window.title("Add New Author")
    
    frame = ttk.Frame(new_author_window, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)
    
    ttk.Label(frame, text="First Name:").grid(row=0, column=0, padx=10, pady=5)
    first_name_entry = ttk.Entry(frame, width=30)
    first_name_entry.grid(row=0, column=1, padx=10, pady=5)
    
    ttk.Label(frame, text="Last Name:").grid(row=1, column=0, padx=10, pady=5)
    last_name_entry = ttk.Entry(frame, width=30)
    last_name_entry.grid(row=1, column=1, padx=10, pady=5)
    
    # Function to handle adding new author
    def submit_author():
        first_name = first_name_entry.get().strip()
        last_name = last_name_entry.get().strip()
        
        if not first_name or not last_name:
            messagebox.showwarning("Missing Information", "Please enter both first and last name.")
            return
        
        add_new_author(first_name, last_name)
        new_author_window.destroy()
    
    ttk.Button(frame, text="Add Author", command=submit_author).grid(row=2, column=0, columnspan=2, pady=10)

    # Function to update author list based on search query

def update_author_list(event=None):
    search_text = entry_search.get().strip().lower()
    all_authors = get_authors()

# Main function to set up the GUI
def main():    
    global entry_title, entry_search, combo_author
    # Declare all globals at the start of the function

    root = tk.Tk()
    root.title("Book Database")
    
    frame = ttk.Frame(root, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)
    
    # Title label and entry
    ttk.Label(frame, text="Title:").grid(row=0, column=0, padx=10, pady=5)
    entry_title = ttk.Entry(frame, width=30)
    entry_title.grid(row=0, column=1, padx=10, pady=5)
    
    ttk.Label(frame, text="Author:").grid(row=1, column=0, padx=10, pady=5)
    combo_author = ttk.Combobox(frame, width=47)
    combo_author.grid(row=1, column=1, padx=10, pady=5)

    # Search Entry
    entry_search = ttk.Entry(frame, width=27)
    entry_search.grid(row=1, column=1, padx=10, pady=5)
    entry_search.bind("<KeyRelease>", update_author_list)  # Bind KeyRelease event here
    
    # Author Combobox
    combo_author = ttk.Combobox(frame, width=27, state="readonly")
    combo_author.grid(row=2, column=1, padx=10, pady=5)
    combo_author.bind("<FocusIn>", lambda event: combo_author.delete(0, "end"))

    
    # Container frame for Add Book and Remove Book buttons
    book_button_frame = ttk.Frame(frame)
    book_button_frame.grid(row=2, column=0, columnspan=2, pady=10,)
    # Add Book Button
    btn_add = ttk.Button(book_button_frame, text="Add Book", command=on_add_book)
    btn_add.pack(side=tk.LEFT, padx=(72, 5))
    # Remove Book Button
    btn_remove = ttk.Button(book_button_frame, text="Remove Book", command=on_remove_book)
    btn_remove.pack(side=tk.LEFT, padx=(5, 10))
    
    # New Author Button
    btn_new_author = ttk.Button(frame, text="New Author", command=add_new_author_window)
    btn_new_author.grid(row=1, column=2, padx=10, pady=5)

    # Remove Author Button
    btn_remove_author = ttk.Button(frame, text="Remove Author", command=on_remove_author)
    btn_remove_author.grid(row=1, column=3, padx=10, pady=5)

    # Adjust column weights to make author combobox, add book, and remove book buttons the same width
    frame.columnconfigure(1, weight=1)  # Author combobox
    frame.columnconfigure(2, weight=1)  # Add Book button
    frame.columnconfigure(3, weight=1)  # Remove Book button

    # Initialize authors list
    #update_author_list()  # Call the function to initialize the author list

    root.mainloop()

    # Function to export database content to Excel

def export_database_to_excel(db_path, excel_file_path):
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Query to fetch data from database tables
        cursor.execute("SELECT * FROM Authors")
        authors_data = cursor.fetchall()

        cursor.execute("SELECT * FROM Books")
        books_data = cursor.fetchall()

        # Check if the workbook exists
        if os.path.exists(excel_file_path):
            wb = load_workbook(excel_file_path)
        else:
            wb = Workbook()

        # Create or select the Authors sheet
        if "Authors" in wb.sheetnames:
            ws_authors = wb["Authors"]
            # Clear existing data (including headers)
            ws_authors.delete_rows(1, ws_authors.max_row)
        else:
            ws_authors = wb.active
            ws_authors.title = "Authors"

        # Create or select the Books sheet
        if "Books" in wb.sheetnames:
            ws_books = wb["Books"]
            # Clear existing data (including headers)
            ws_books.delete_rows(1, ws_books.max_row)
        else:
            ws_books = wb.create_sheet(title="Books")

        # Write headers for Authors and Books sheets
        ws_authors.append(["author_id", "author_name"])
        ws_books.append(["book_id", "title", "author_id"])

        # Write data to Authors sheet
        for author in authors_data:
            ws_authors.append(author)

        # Write data to Books sheet
        for book in books_data:
            ws_books.append(book)

        # Save the workbook
        wb.save(excel_file_path)
        print(f"Excel file saved successfully: {excel_file_path}")

    except sqlite3.Error as e:
        print(f"Error exporting database to Excel: {e}")
    finally:
        if conn:
            conn.close()

# Entry point of the script
if __name__ == "__main__":
    print(f"Script directory: {script_dir}")
    print(f"Database path: {db_path}")
    print(f"SQL file path: {sql_file_path}")
    print(f"Excel file path: {excel_file_path}")

    export_database_to_excel(db_path, excel_file_path)

    # Ensure the database and tables are set up
    if not os.path.exists(db_path):
        execute_sql_script(sql_file_path)
        print("SQL script executed successfully.")

    main()

