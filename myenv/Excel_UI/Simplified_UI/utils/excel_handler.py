import openpyxl
import logging
import os
import json
from datetime import datetime, timedelta
from tkinter import messagebox
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font

#from scripts.config_handler import load_config, save_config

CONFIG_FILE = os.path.join(os.path.dirname(__file__), "config.json")


# Configure logging
logging.basicConfig(
    filename="error_log.txt",
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s",
)

logging.info(f"Resolved CONFIG_FILE path: {CONFIG_FILE}")

def extend_date_row(workbook, sheet, file_path, start_column=3):

    #wb = openpyxl.load_workbook(file_path)
    #ws = wb[sheet_name]
    """Ensure all missing dates are added to the date row up until today, formatted as dd-mmm."""
    try:
        today = datetime.now().date()
        print(f"DEBUG: Extending date row up to {today}")
        print(f"DEBUG: Sheet title -> {sheet.title}")
        print(f"DEBUG: Current max column -> {sheet.max_column}")

        # Step 1: Find last valid date in row 1
        last_date = None
        last_column = start_column - 1

        for col in range(start_column, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            parsed_date = None

            if cell_value:
                if isinstance(cell_value, datetime):
                    parsed_date = cell_value.date()
                elif isinstance(cell_value, str):
                    try:
                        parsed_date = datetime.strptime(cell_value, "%d-%b").date()
                    except ValueError:
                        try:
                            parsed_date = datetime.strptime(cell_value, "%m/%d/%Y").date()
                        except ValueError:
                            parsed_date = None

            if parsed_date:
                print(f"DEBUG: Found existing date in col {col} -> {parsed_date}")
                last_date = parsed_date
                last_column = col

        if not last_date:
            last_date = datetime(today.year, 1, 1).date()
            print(f"DEBUG: No existing date found. Defaulting to: {last_date}")

        # Step 2: Append missing dates
        next_date = last_date + timedelta(days=1)
        current_column = last_column + 1
        print(f"DEBUG: Starting extension from {next_date}")

        while next_date <= today:
            # Confirm it's not already in the sheet
            is_duplicate = False
            for col in range(start_column, sheet.max_column + 1):
                val = sheet.cell(row=1, column=col).value
                if isinstance(val, datetime) and val.date() == next_date:
                    is_duplicate = True
                    break

            if not is_duplicate:
                cell = sheet.cell(row=1, column=current_column)
                cell.value = next_date
                cell.number_format = "DD-MMM"
                print(f"DEBUG: Added missing date -> {next_date.strftime('%d-%b')} at column {current_column}")
                current_column += 1

            next_date += timedelta(days=1)

        print(f"DEBUG: Completed extension. Final max column: {sheet.max_column}")
        print("DEBUG: Final date row after extension:")
        
        # Save the workbook after adding missing dates
        try:
            print("DEBUG: Saving workbook after extending date row...")
            workbook.save(file_path)
            print(f"DEBUG: Workbook saved to {file_path}")
        except Exception as e:
            print(f"ERROR: Failed to save workbook after extending date row: {e}")
            logging.error(f"Failed to save workbook: {e}")

        
        for col in range(1, sheet.max_column + 1):
            cell_val = sheet.cell(row=1, column=col).value
            print(f"  Col {col}: {cell_val} ({type(cell_val)})")

    except Exception as e:
        error_message = f"Exception in extend_date_row: {e}"
        print(f"ERROR: {error_message}")
        logging.error(error_message)
        messagebox.showerror("Error", error_message)    

def save_days_without_incident_data(counter, last_date):
    """Save the Days without Incident data to a JSON file."""
    try:
        config = load_config()
        config["counter"] = counter
        config["last_date"] = last_date.strftime("%Y-%m-%d")
        save_config(config)
        logging.info(f"Days without Incident updated: {config}")
    except Exception as e:
        error_message = f"Error saving Days without Incident data: {e}"
        logging.error(error_message)
        messagebox.showerror("Error", error_message)

def load_days_without_incident_data():
    """Load the Days without Incident data from a JSON file."""
    if not CONFIG_FILE:
        raise ValueError("CONFIG_FILE is not set.")
    if not os.path.exists(CONFIG_FILE):
        return {"counter": 0, "last_date": datetime.now().strftime("%Y-%m-%d")}
    try:
        with open(CONFIG_FILE, "r") as f:
            return json.load(f)
    except Exception as e:
        logging.error(f"Error loading Days without Incident data: {e}")
        raise

def find_or_add_date_column(sheet, selected_date, start_column=3):
    """Find the column for the selected date without adding a new one."""
    try:
        for col in range(start_column, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            
            # Convert cell_value to a datetime object if possible
            cell_date = None
            if isinstance(cell_value, datetime):
                cell_date = cell_value.date()
            elif isinstance(cell_value, str):
                try:
                    # First attempt: "dd-MMM" format (e.g., "04-Mar")
                    cell_date = datetime.strptime(cell_value, "%d-%b").date()
                except ValueError:
                    try:
                        # Second attempt: "MM/DD/YYYY" format
                        cell_date = datetime.strptime(cell_value, "%m/%d/%Y").date()
                    except ValueError:
                        continue  # Skip non-date strings

            if cell_date == selected_date:
                return col  # Found the correct column

        logging.error(f"Date {selected_date.strftime('%d-%b')} not found in the sheet.")
        return None  # Return None if not found
    except Exception as e:
        logging.error(f"Error finding date column: {e}")
        return None

def load_workbook(file_path):
    """Load an Excel workbook."""
    try:
        return openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        logging.error(f"File not found: {file_path}")
        raise FileNotFoundError(f"The file {file_path} does not exist.")
    except Exception as e:
        logging.error(f"Failed to load workbook: {file_path} - {e}")
        raise

def save_workbook(workbook, file_path):
    """Save the workbook to the specified file path."""
    try:
        print(f"Attempting to save workbook to: {file_path}")
        workbook.save(file_path)
        logging.info(f"Workbook saved to {file_path}")
    except Exception as e:
        logging.error(f"Failed to save workbook: {e}")
        print(f"DEBUG: Failed to save workbook: {e}")
        #raise

def get_sheet(workbook, sheet_name):
    """Retrieve a specific sheet by name."""
    try:
        # Debug: Log available sheets
        available_sheets = list(workbook.sheetnames)
        print(f"get_sheet called. Available sheets: {available_sheets}")

        return workbook[sheet_name]
    except KeyError:
        logging.error(
            f"Sheet not found: {sheet_name}. Available sheets: {available_sheets}")
        print(
            f"Sheet not found: {sheet_name}. Available sheets: {available_sheets}")
        raise KeyError(
            f"The sheet {sheet_name} does not exist in the workbook.")

def write_to_cell(sheet, row, col, value):
    """Write a value to a specific cell in the sheet."""
    try:
        sheet.cell(row=row, column=col).value = value
        print(
            f"Value '{value}' written to Row {row}, Column {col} in Sheet '{sheet.title}'")
    except Exception as e:
        print(f"Error writing to cell ({row}, {col}): {e}")
        raise

def read_cell(sheet, row, col):
    """Read a value from a specific cell."""
    try:
        return sheet.cell(row=row, column=col).value
    except Exception as e:
        logging.error(f"Failed to read cell: row={row}, col={col} - {e}")
        raise

def load_config():
    """Load configuration data from the JSON file, creating default values if necessary."""
    try:
        if not os.path.exists(CONFIG_FILE):
            logging.warning(f"Config file not found. Creating a new one at {CONFIG_FILE}.")
            default_config = {
                "window_geometry": "800x600",
                "counter": 0,
                "last_date": datetime.now().strftime("%Y-%m-%d"),
                "last_used_workbook": "",  # ✅ Ensure this key exists
                "boxing_tier_file": "",
                "boxing_log_file": ""
            }
            save_config(default_config)  # Save default config
            return default_config

        with open(CONFIG_FILE, "r") as f:
            config = json.load(f)

        # ✅ Ensure all required keys exist (to prevent issues)
        default_keys = {
            "window_geometry": "800x600",
            "counter": 0,
            "last_date": datetime.now().strftime("%Y-%m-%d"),
            "last_used_workbook": "",
            "boxing_tier_file": "",
            "boxing_log_file": "",
            "last_selected_workbook": "Boxing Tier"
        }

        for key, default_value in default_keys.items():
            if key not in config:
                config[key] = default_value  # Add missing keys

        save_config(config)  # Save the updated config with missing values
        return config

    except json.JSONDecodeError:
        logging.error("Corrupt config.json detected! Resetting to default settings.")
        print("DEBUG: config.json is corrupted. Resetting to default.")

        default_config = {
            "window_geometry": "800x600",
            "counter": 0,
            "last_date": datetime.now().strftime("%Y-%m-%d"),
            "last_used_workbook": "",
            "boxing_tier_file": "",
            "boxing_log_file": ""
        }

        save_config(default_config)  # Save default config
        return default_config

    except Exception as e:
        logging.error(f"Error loading config: {e}")
        return {}

def save_config(data):
    """Save configuration data to the JSON file."""
    try:
        os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)  # Ensure the directory exists
        with open(CONFIG_FILE, "w") as f:
            json.dump(data, f, indent=4)  # Save in a readable format
        logging.info(f"Config saved to {CONFIG_FILE}")
    except Exception as e:
        logging.error(f"Error saving config: {e}")

def save_last_file_path(file_path):
    """Save the last selected file path to the JSON configuration file."""
    try:
        config = load_config()
        config["last_used_workbook"] = file_path
        save_config(config)
        logging.info(f"Saved last used workbook: {file_path}")
        print(f"DEBUG: Last used workbook saved -> {file_path}")
    except Exception as e:
        logging.error(f"Failed to save last file path: {e}")

def load_last_file_path():
    """Load the last selected file path from the configuration file."""
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r") as f:
                data = json.load(f)
                return data.get("last_file")
    except Exception as e:
        logging.error(f"Failed to load last file path: {e}")
    return None

#def load_last_file_path():
    """Load the last selected file path from the JSON configuration file."""
    try:
        config = load_config()
        return config.get("last_used_workbook")
    except Exception as e:
        logging.error(f"Failed to load last file path: {e}")
        return None

def calculate_truck_fill_percentage(value):
    """
    Validate and calculate the Truck Fill % based on the maximum value of 26.
    :param value: The entered value as a number (float or int).
    :return: The calculated percentage as a string with a '%' symbol.
    :raises ValueError: If the value is not a number or is outside the valid range.
    """
    try:
        value = float(value)  # Ensure the value is numeric
        if 0 <= value <= 26:
            percentage = (value / 26) * 100  # Calculate percentage
            return f"{percentage:.2f}%"  # Format with the '%' symbol
        else:
            raise ValueError("Value must be between 0 and 26.")
    except ValueError:
        raise ValueError(
            "Invalid input. Please enter a numeric value between 0 and 26.")

def create_or_update_dashboard(workbook, data_sheet, file_path, selected_date, start_column=3):
    try:
        # Create sheet if it doesn't exist
        if "Dashboard (Implemented)" in workbook.sheetnames:
            sheet = workbook["Dashboard (Implemented)"]
        else:
            sheet = workbook.create_sheet("Dashboard (Implemented)")
            print("DEBUG: Created 'Dashboard (Implemented)' sheet.")
            build_dashboard_layout(sheet)

        # --- Get the correct date column from the data sheet ---
        date_col = find_or_add_date_column(data_sheet, selected_date, start_column)
        if date_col is None:
            raise ValueError(f"Date {selected_date.strftime('%d-%b')} not found.")

        # --- Define mappings from dashboard rows to data sheet rows ---
        field_map = {
            "Days Without Incident": (2, 2),
            "Haz ID's": (3, 2),
            "Safety Gemba Walk": (4, 2),
            "7S (Zone 26)": (5, 2),
            "7S (Zone 51)": (6, 2),
            "Errors": (8, 2),
            "PCD Returns": (9, 2),
            "Jobs on Hold": (11, 2),
            "Productivity": (12, 2),
            "OTIF": (13, 2),
            "Truck Fill %": (14, 2),
            "Recognitions": (15, 2),
            "MC Compliance": (16, 2),
            "Cost Savings": (18, 2),
            "Rever's": (19, 2),
            "Project's": (20, 2),
        }

        # --- Insert values into dashboard ---
        for label, (dash_row, dash_col) in field_map.items():
            for r in range(2, data_sheet.max_row + 1):
                if data_sheet.cell(row=r, column=1).value == label:
                    value = data_sheet.cell(row=r, column=date_col).value
                    sheet.cell(row=dash_row, column=dash_col).value = value
                    break

        workbook.save(file_path)
        print("DEBUG: Dashboard values updated.")

    except Exception as e:
        logging.error(f"Dashboard update failed: {e}")
        messagebox.showerror("Error", f"Dashboard update failed: {e}")

def build_dashboard_layout(sheet):
    # Add static header layout with merged cells and colors
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)

    # Safety Section
    sheet.merge_cells("B1:G1")
    sheet["B1"] = "Safety"
    sheet["B1"].fill = yellow_fill
    sheet["B1"].alignment = center_align
    sheet["B1"].font = bold_font

    sheet["B2"] = "Days Without Incident"
    sheet["B3"] = "Haz ID's"
    sheet["B4"] = "Safety Gemba Walk"
    sheet["B5"] = "7S (Zone 26)"
    sheet["B6"] = "7S (Zone 51)"

    # Quality Section
    sheet.merge_cells("B8:D8")
    sheet["B8"] = "Quality"
    sheet["B8"].fill = PatternFill(start_color="00FFCC", end_color="00FFCC", fill_type="solid")
    sheet["B8"].alignment = center_align
    sheet["B8"].font = bold_font

    sheet["B9"] = "Errors"
    sheet["B10"] = "PCD Returns"

    # Operations Section
    sheet.merge_cells("B12:E12")
    sheet["B12"] = "Operations"
    sheet["B12"].fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")
    sheet["B12"].alignment = center_align
    sheet["B12"].font = bold_font

    sheet["B13"] = "Jobs on Hold"
    sheet["B14"] = "Productivity"
    sheet["B15"] = "OTIF"
    sheet["B16"] = "Truck Fill %"

    # People Section
    sheet.merge_cells("B18:D18")
    sheet["B18"] = "People"
    sheet["B18"].fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    sheet["B18"].alignment = center_align
    sheet["B18"].font = bold_font

    sheet["B19"] = "Recognitions"
    sheet["B20"] = "MC Compliance"

    # Extras (PPI, Cost, etc.) could be added here similarly

    # Autofit columns
    for col in range(2, 6):
        sheet.column_dimensions[get_column_letter(col)].width = 25