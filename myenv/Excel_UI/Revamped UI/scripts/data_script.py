import pandas as pd
from openpyxl import load_workbook

def update_data_sheet(file_path, new_data, date):
    # Load workbook and select the 'Data' sheet
    wb = load_workbook(file_path)
    ws = wb['Data']
    
    # Check for the correct date column
    date_col = None
    for cell in ws[1]:
        if cell.value == date:
            date_col = cell.column

    if not date_col:
        raise ValueError(f"Date {date} not found in Data sheet.")

    # Write new data to the appropriate rows
    row_mapping = {
        'Days without Incident': 2,
        'Haz ID\'s': 3,
        'Safety Gemba Walk': 4,
        '7S (Zone 26)': 5,
        '7S (Zone 51)': 6,
        'Errors': 7,
        'PCD Returns': 8,
        'Jobs on Hold': 9,
        'Productivity': 10,
        'OTIF %': 11,
        'Huddles': 12,
        'Truck Fill %': 13,
        'Recognitions': 14,
        'MC Compliance': 15,
        'Cost Savings': 16,
        'Rever\'s': 17,
        'Project\'s': 18
    }

    for key, value in new_data.items():
        if key in row_mapping:
            ws.cell(row=row_mapping[key], column=date_col, value=value)
    
    # Save workbook after updating
    wb.save(file_path)
    print(f"Data sheet updated for date {date}")
