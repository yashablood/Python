from openpyxl import load_workbook

def append_dashboard(file_path, dashboard_data):
    wb = load_workbook(file_path)
    ws = wb['Dashboard']

    # Find the first empty row
    row = ws.max_row + 1

    # Assuming recognition_data is a dictionary with column names and values
    for col_num, (col_name, value) in enumerate(dashboard_data.items(), start=1):
        ws.cell(row=row, column=col_num, value=value)

    wb.save(file_path)
    print("New dashboard data appended.")
