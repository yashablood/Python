import openpyxl

def add_data_to_data_sheet(workbook, sheet_name, data_dict, date_column):
    sheet = workbook[sheet_name]
    
    # Find the correct column for the date
    for col in range(2, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == date_column:
            target_column = col
            break
    else:
        raise ValueError(f"Date {date_column} not found in the first row.")
    
    # Define row mappings for each data type
    row_mappings = {
        "Days without Incident": 2,
        "Haz ID's": 3,
        "Safety Gemba Walk": 4,
        "7S (Zone 26)": 5,
        "7S (Zone 51)": 6,
        "Errors": 7,
        "PCD Returns": 8,
        "Jobs on Hold": 9,
        "Productivity": 10,
        "OTIF %": 11,
        "Huddles": 12,
        "Truck Fill %": 13,
        "Recognitions": 14,
        "MC Compliance": 15,
        "Cost Savings": 16,
        "Rever's": 17,
        "Project's": 18
    }

    # Add data to the corresponding rows
    for key, value in data_dict.items():
        row = row_mappings.get(key)
        if row:
            sheet.cell(row=row, column=target_column).value = value

# Example usage:
# data_to_add = {
#     "Days without Incident": 5,
#     "Haz ID's": 2,
#     "Safety Gemba Walk": 1,
#     "Errors": 0,
# }
# add_data_to_data_sheet(workbook, "Data", data_to_add, "2023-10-09")
