import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import pandas as pd
import openpyxl 
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import tkcalendar
from tkcalendar import DateEntry


# Global variables
file_path = None
current_sheet_name = None
entries = {}

script_dir = os.path.dirname(__file__)

# Load the Excel file
file_path = 'Boxing Tier.xlsx'
print("File path:", os.path.abspath(file_path))

# Specify the worksheet name where the data resides
def process_file(file_path):
    try:    
        wb = load_workbook(filename=file_path)
        # Optionally save all the sheetnames to global variable for later use.
        worksheet_names = wb.sheetnames

        for worksheet_name in worksheet_names:
            # Read data from the worksheet
            df = pd.read_excel(file_path, sheet_name=worksheet_name,)
            ws = wb[worksheet_name]

            # Read required sheets
            Dashboard_Rev_2_df = pd.read_excel(file_path, sheet_name='Dashboard Rev 2')
            Data_df = pd.read_excel(file_path, sheet_name='Data')
            Recognitions_df = pd.read_excel(file_path, sheet_name='Recognitions')
            Error_Tracker_df = pd.read_excel(file_path, sheet_name='Error Tracker')
            Production_df = pd.read_excel(file_path, sheet_name='Production')
            OTIF_df = pd.read_excel(file_path, sheet_name='OTIF')

        # Save the changes to the Excel file
        wb.save(file_path)
        return "Processing complete!"
    except Exception as e:
        return f"An error occurred: {e}"

# Select a file and update the dropdowns
def select_file():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        file_path_label.config(text=os.path.basename(file_path))
        result = process_file(file_path)
        update_dropdowns(file_path)
        result_text.insert(tk.END, result + "\n")

def update_dropdowns(file_path):
    wb = load_workbook(filename=file_path)
    worksheet_names = wb.sheetnames
    sheet_name_combobox['values'] = worksheet_names
    sheet_name_combobox.current(0)
    #update_column_names()
    update_sheet_window()

def update_sheet_window(event=None):
    global current_sheet_name
    sheet_name = sheet_name_combobox.get()
    current_sheet_name = sheet_name

    # Destroy previous widgets in the sheet window
    for widget in sheet_window_frame.winfo_children():
        widget.destroy()

    # Determine which sheet-specific function to call based on sheet_name
    if sheet_name == 'Dashboard Rev 2':
        create_dashboard_rev_2_window()
    elif sheet_name == 'Data':
        create_data_window()
    elif sheet_name == 'Recognitions':
        create_recognitions_window()
    elif sheet_name == 'Error Tracker':
        create_error_tracker_window()
    elif sheet_name == 'Production':
        create_production_window()
    elif sheet_name == 'OTIF':
        create_otif_window()
    else:
        # Handle unknown sheet name
        pass


def create_dashboard_rev_2_window():
    # Define specific widgets and layout for 'Dashboard Rev 2' sheet
    pass

def create_data_window():
    global file_path

    # Destroy previous widgets in the sheet window
    for widget in sheet_window_frame.winfo_children():
        widget.destroy()

    # Read the Excel file without headers
    try:
        Data_df_no_header = pd.read_excel(file_path, sheet_name='Data', header=None)
    except FileNotFoundError:
        messagebox.showerror("Error", f"File not found: {file_path}")
        return
    except pd.errors.EmptyDataError:
        messagebox.showerror("Error", f"File is empty: {file_path}")
        return

    # Read the Excel file with headers to get date columns correctly
    try:
        Data_df_with_header = pd.read_excel(file_path, sheet_name='Data')
    except FileNotFoundError:
        messagebox.showerror("Error", f"File not found: {file_path}")
        return
    except pd.errors.EmptyDataError:
        messagebox.showerror("Error", f"File is empty: {file_path}")
        return

    # Assuming row names (labels) are in column B, rows 2 to 18 (1-based indexing, thus 1:17 in 0-based indexing)
    row_names = Data_df_no_header.iloc[1:18, 1].tolist()

    print("Date DataFrame:")
    print(Data_df_no_header.iloc[:, 1].tolist())

    tk.Label(sheet_window_frame, text="Data Sheet", font=("Arial", 14)).grid(row=0, column=0, columnspan=3, pady=(0, 10))

    # Dynamically create labels and entry fields for each row
    entry_fields = {}
    start_row = 1  # Start row index for entries
    column_index = 0  # Fixed column index for labels

    for i, row_name in enumerate(row_names):
        # Calculate row index for each label and entry field
        row_index = start_row + i

        # Create label and entry field
        tk.Label(sheet_window_frame, text=f"{row_name}:").grid(row=row_index, column=column_index, sticky=tk.W, pady=(5, 5))
        
        entry_fields[row_name] = tk.Entry(sheet_window_frame)
        entry_fields[row_name].grid(row=row_index, column=column_index + 1, columnspan=1, sticky=tk.W + tk.E, pady=(5, 5))

    # Add DateEntry widget for date selection
    tk.Label(sheet_window_frame, text="Select Date:").grid(row=start_row + len(row_names), column=0, sticky=tk.W, pady=(5, 5))
    date_entry = DateEntry(sheet_window_frame, date_pattern="MM/dd/yyyy")
    date_entry.grid(row=start_row + len(row_names), column=1, sticky=tk.W + tk.E, pady=(5, 5))

    # Add a submit button to save data
    submit_button = tk.Button(sheet_window_frame, text="Submit", command=lambda: save_data(entry_fields, date_entry.get_date(), Data_df_no_header, Data_df_with_header))
    submit_button.grid(row=start_row + len(row_names) + 1, column=0, columnspan=2, pady=(10, 0))

    # Ensure the main window updates correctly after adding widgets
    sheet_window_frame.update_idletasks()

def save_data(entry_fields, selected_date, Data_df_no_header, Data_df_with_header):
    global file_path

    try:
        # Format the date
        formatted_date = selected_date.strftime("%m/%d/%Y")

        # Ensure Data_df_with_header only contains relevant data columns (excluding non-date columns)
        data_columns = Data_df_with_header.columns[2:]  # Assuming the first two columns are non-date columns

        # Format the date columns into the desired format
        formatted_columns = pd.to_datetime(data_columns, format="%m/%d/%Y").strftime("%m/%d/%Y")

        # Check if the selected date exists in formatted columns
        if formatted_date not in formatted_columns:
            messagebox.showerror("Error", f"Date {formatted_date} not found in the sheet.")
            return
            
        # Assign formatted columns to Data_df_no_header
        Data_df_no_header.columns = list(Data_df_with_header.columns[:2]) + list(formatted_columns)

        # Append data to the correct column
        for row_name, entry in entry_fields.items():
            value = entry.get()
            if value:  # If there's a value entered
                # Locate the correct row based on the label in column B
                row_index = Data_df_no_header[Data_df_no_header.iloc[:, 1] == row_name].index[0]
                Data_df_no_header.loc[row_index, formatted_date] = value


            # Check if we're updating "Truck Fill %" and the value is a valid number
            if row_name == "Truck Fill %" and value.isdigit():
                fill_value = min(int(value), 26)  # Cap the value at 26
                percentage = (fill_value / 26)
                
                # Find the index of "Truck Fill %" in column B
                truck_fill_row = Data_df_no_header[Data_df_no_header.iloc[:, 1].str.strip() == "Truck Fill %"].index[0]

                # Update the truck fill percentage in the appropriate column
                Data_df_no_header.loc[truck_fill_row, formatted_date] = percentage

            elif value:  # If there's a value entered for other rows
                # Locate the correct row based on the label in column B
                row_index = Data_df_no_header[Data_df_no_header.iloc[:, 1].str.strip() == row_name].index[0]
                Data_df_no_header.loc[row_index, formatted_date] = value


        # Filter the DataFrame to only include rows 2 to 18
        Data_df_to_save = Data_df_no_header.iloc[1:18]

        # Save the updated DataFrame back to the Excel file, excluding the first row
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            Data_df_to_save.to_excel(writer, sheet_name='Data', index=False, header=False, startrow=1)  # Start writing from row 2

        messagebox.showinfo("Success", "Data has been appended successfully!")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")
    
    pass

def create_recognitions_window():
    global file_path

    # Destroy previous widgets in the sheet window
    for widget in sheet_window_frame.winfo_children():
        widget.destroy()

    Recognitions_df = pd.read_excel(file_path, sheet_name='Recognitions')
    column_names = Recognitions_df.columns

    tk.Label(sheet_window_frame, text="Recognitions Sheet", font=("Arial", 14)).grid(row=0, column=0, columnspan=2, pady=(0, 10))

    # Dynamically create labels and entry fields for each column
    entry_fields = {}
    row_index = 1  # Start row index for entries
    for column_name in column_names:
        tk.Label(sheet_window_frame, text=f"{column_name}:").grid(row=row_index, column=0, sticky=tk.W, pady=(5, 5))

        if column_name.lower() == "date":  # Check for "Date" column
            # DateEntry widget for date and time input (military time)
            entry_fields[column_name] = DateEntry(sheet_window_frame, date_pattern="dd-mm-yyyy")
            entry_fields[column_name].grid(row=row_index, column=1, sticky=tk.W, pady=(5, 5))

            # Button to add current date
            add_date_button = tk.Button(sheet_window_frame, text="Add Current Date",
                command=lambda field=entry_fields[column_name]: add_current_date(field))
            add_date_button.grid(row=row_index, column=2, sticky=tk.W, padx=(10, 0))  # Add some padding

        else:
            entry_fields[column_name] = tk.Entry(sheet_window_frame)
            entry_fields[column_name].grid(row=row_index, column=1, columnspan=2, sticky=tk.W + tk.E, pady=(5, 5))

        row_index += 1  # Move to the next row for the next entry

    # Function to submit recognition data
    def submit_recognition():
        global file_path
        sheet_name = 'Recognitions'
        Recognitions_df = pd.read_excel(file_path, sheet_name=sheet_name)
        column_names = Recognitions_df.columns

        data = {}
        for column_name in column_names:
            if isinstance(entry_fields[column_name], DateEntry):
                # Format date in "dd-mmm-yyyy" format
                data[column_name] = entry_fields[column_name].get_date().strftime("%d-%b-%Y")
            else:
                data[column_name] = entry_fields[column_name].get().strip()

        if not all(data.values()):
            messagebox.showwarning("Warning", "Please fill in all fields.")
            return

        result = append_data(file_path, "Recognitions", data)
        result_text.insert(tk.END, result + "\n")

    # Button to submit recognition data
    submit_button = tk.Button(sheet_window_frame, text="Submit Recognition", command=submit_recognition)
    submit_button.grid(row=row_index, column=0, columnspan=3, pady=10)

    # Ensure the main window updates correctly after adding widgets
    sheet_window_frame.update_idletasks()

    pass

def create_error_tracker_window():
    # Define specific widgets and layout for 'Error Tracker' sheet

    global file_path

    # Destroy previous widgets in the sheet window
    for widget in sheet_window_frame.winfo_children():
        widget.destroy()

    Error_Tracker_df = pd.read_excel(file_path, sheet_name='Error Tracker')
    column_names = Error_Tracker_df.columns

    tk.Label(sheet_window_frame, text="Error Tracker Sheet", font=("Arial", 14)).grid(row=0, column=0, columnspan=2, pady=(0, 10))

    # Dynamically create labels and entry fields for each column
    entry_fields = {}
    row_index = 1  # Start row index for entries
    for column_name in column_names:
        tk.Label(sheet_window_frame, text=f"{column_name}:").grid(row=row_index, column=0, sticky=tk.W, pady=(5, 5))

        if column_name.lower() == "date":  # Check for "Date" column
            # DateEntry widget for date and time input (military time)
            entry_fields[column_name] = DateEntry(sheet_window_frame, date_pattern="dd-mm-yyyy")
            entry_fields[column_name].grid(row=row_index, column=1, sticky=tk.W, pady=(5, 5))

            # Button to add current date
            add_date_button = tk.Button(sheet_window_frame, text="Add Current Date",
                command=lambda field=entry_fields[column_name]: add_current_date(field))
            add_date_button.grid(row=row_index, column=2, sticky=tk.W, padx=(10, 0))  # Add some padding

        else:
            entry_fields[column_name] = tk.Entry(sheet_window_frame)
            entry_fields[column_name].grid(row=row_index, column=1, columnspan=2, sticky=tk.W + tk.E, pady=(5, 5))

        row_index += 1  # Move to the next row for the next entry

    # Function to submit error data
    def submit_error():
        global file_path
        sheet_name = 'Error Tracker'
        Error_Tracker_df = pd.read_excel(file_path, sheet_name=sheet_name)
        column_names = Error_Tracker_df.columns

        data = {}
        for column_name in column_names:
            if isinstance(entry_fields[column_name], DateEntry):
                # Format date in "dd-mmm-yyyy" format
                data[column_name] = entry_fields[column_name].get_date().strftime("%d-%b-%Y")
            else:
                data[column_name] = entry_fields[column_name].get().strip()

        if not all(data.values()):
            messagebox.showwarning("Warning", "Please fill in all fields.")
            return

        result = append_data(file_path, "Error Tracker", data)
        result_text.insert(tk.END, result + "\n")

    # Button to submit error data
    submit_button = tk.Button(sheet_window_frame, text="Submit Error", command=submit_error)
    submit_button.grid(row=row_index, column=0, columnspan=3, pady=10)

    # Ensure the main window updates correctly after adding widgets
    sheet_window_frame.update_idletasks()


    pass

def create_production_window():
    # Define specific widgets and layout for 'Production' sheet
    pass

def create_otif_window():
    # Define specific widgets and layout for 'OTIF' sheet
    pass

    

def append_data_to_file():
    global file_path
    if not file_path:
        messagebox.showwarning("Warning", "Please select a file first.")
        return

    sheet_name = sheet_name_combobox.get()
    if not sheet_name:
        messagebox.showwarning("Warning", "Please select a sheet name.")
        return

    # Validate if entries have been initialized
    if not entries:
        messagebox.showwarning("Warning", "No data fields found.")
        return
        
    data = {}
    for column, entry in entries.items():
        if isinstance(entry, tk.Text):
            data[column] = entry.get("1.0", tk.END).strip()
        else:
            data[column] = entry.get().strip()


    # Validate if any of the data fields are empty
    if not data[column]:
        messagebox.showwarning("Warning", f"Please fill in the '{column}' field.")
        return

    result = append_data(file_path, sheet_name, data)
    result_text.insert(tk.END, result + "\n")

def add_current_date(date_field):
    import datetime
    # Set the current date in the DateEntry field
    date_field.set_date(datetime.date.today())

def append_data(file_path, sheet_name, data):
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        # Get the header row (first row)
        header_row = ws[1]

        # Initialize dictionary to map column names to their indices
        column_index = {cell.value: cell.column for cell in header_row if cell.value}

        # Determine the last row with data in a specific column (e.g., Column A)
        last_row = ws.max_row

        # Ensure that the last row is correctly identified (ignoring empty rows)
        while last_row > 0 and all(cell.value is None for cell in ws[last_row]):
            last_row -= 1

        # Increment last_row to get the next empty row
        new_row = last_row + 1

        # Iterate data in the new row
        for column, value in data.items():
            col_index = column_index.get(column)
            cell = ws.cell(row=new_row, column=col_index, value=value)
            ws.cell(row=new_row, column=col_index, value=value)

            # Check if the column is the date column and align the cell to the right
            if column.lower() == "date":  # Adjust this condition to match your date column name
                cell.alignment = Alignment(horizontal='right')

        wb.save(file_path)
        return "Data appended successfully!"
    except Exception as e:
        return f"An error occurred: {e}"

# Function to update the canvas scroll region and center content
def update_canvas(event=None):
    # Update the scroll region of the canvas
    canvas.configure(scrollregion=canvas.bbox("all"))
    
    # Center the content
    canvas_width = canvas.winfo_width()
    content_width = sheet_window_frame.winfo_width()
    
    # Calculate x offset to center the content
    x_offset = (canvas_width - content_width) // 2 if content_width < canvas_width else 0
    
    # Adjust position of the window (frame)
    canvas.itemconfig(window, x=x_offset)


# Create the main window
root = tk.Tk()
root.title("Sample Data Processor")

# Create a frame for the file selection
frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

# Add a label and button for file selection
file_path_label = tk.Label(frame, text="No file selected")
file_path_label.pack(anchor="w")

select_file_button = tk.Button(frame, text="Select File", command=select_file)
select_file_button.pack()

# Add entry fields for appending data
entry_frame = tk.Frame(root)
entry_frame.pack(padx=10, pady=10)

tk.Label(frame, text="Sheet Name:").pack()
sheet_name_combobox = ttk.Combobox(frame, state="readonly")
sheet_name_combobox.pack()

# Frame for the canvas and scrollbar
canvas_frame = tk.Frame(root)
canvas_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

# Create the Canvas and Scrollbar
canvas = tk.Canvas(canvas_frame)
scrollbar = tk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=canvas.yview)
canvas.configure(yscrollcommand=scrollbar.set)

# Pack the Scrollbar and Canvas
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Create a frame inside the canvas to hold the content
sheet_window_frame = tk.Frame(canvas)
window = canvas.create_window(0, 0, window=sheet_window_frame, anchor='nw')

# Add the frame to the canvas and store the window object globally
window = canvas.create_window(0, 0, window=sheet_window_frame, anchor="center")

# Bind to the canvas resize event
canvas.bind("<Configure>", update_canvas)

# Configure the canvas scroll region
def update_canvas(event=None):
    canvas.configure(scrollregion=canvas.bbox("all"))

# Frame for sheet-specific widgets (dynamic frame)
dynamic_frame = tk.Frame(root)
dynamic_frame.pack(padx=10, pady=10)

# Add a text widget for showing results
result_text = tk.Text(root, height=10, width=80)
result_text.pack(padx=10, pady=10)

# Start the GUI event loop
root.mainloop()