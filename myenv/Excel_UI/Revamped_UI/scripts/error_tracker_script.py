from openpyxl import load_workbook

def append_error_tracker(file_path, error_tracker_data):
    wb = load_workbook(file_path)
    ws = wb['Error Tracker']

    # Find the first empty row
    row = ws.max_row + 1

    # Assuming recognition_data is a dictionary with column names and values
    for col_num, (col_name, value) in enumerate(error_tracker_data.items(), start=1):
        ws.cell(row=row, column=col_num, value=value)

    wb.save(file_path)
    print("New error tracker data appended.")
