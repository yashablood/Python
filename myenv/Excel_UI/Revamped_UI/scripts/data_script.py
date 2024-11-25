from openpyxl import load_workbook

def update_data_sheet(file_path, data, date):
    # Load workbook and check sheets
    print(f"Opening file: {file_path}")
    workbook = load_workbook(file_path)
    if "Data" not in workbook.sheetnames:
        print("Sheet 'Data' not found.")
        return  # Exit if sheet is missing
    
    sheet = workbook["Data"]

    # Locate the column corresponding to the selected date
    date_column = None
    for cell in sheet[1]:  # Assumes dates are in the first row (starting at C1)
        if cell.value == date:
            date_column = cell.column
            break

    if date_column is None:
        print(f"Date {date} not found in {sheet} sheet.")
        return

    # Update rows based on labels in data
    row_mapping = {
        "Days without Incident": 2,
        "Haz ID's": 3,
        "Safety Gemba Walk": 4,
        "7S (Zone 26)": 5,
        "7S (Zone 51)": 6,
        "Errors": 7,
        "PCD Returns": 8,
        "Jobs on Hold": 9,
        "Productivity": 10,
        "OTIF %": 11,
        "Huddles": 12,
        "Truck Fill %": 13,
        "Recognitions": 14,
        "MC Compliance": 15,
        "Cost Savings": 16,
        "Rever's": 17,
        "Project's": 18
    }

    # Write data to the appropriate cells
    for label, value in data.items():
        if label in row_mapping:
            row = row_mapping[label]
            print(f"Writing {value} to {sheet.cell(row=row, column=date_column).coordinate}")
            sheet.cell(row=row, column=date_column).value = value

    # Save changes to the file
    print("Trying to save")
    workbook.save(file_path)
    print(f"Data sheet updated successfully for date {date}.")
