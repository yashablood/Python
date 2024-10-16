from openpyxl import load_workbook

def load_data(file_path):
    # Load data from the specified Excel file
    print("Recognitions script loaded")
    pass

def append_recognitions(file_path, recognition_data):
    wb = load_workbook(file_path)
    ws = wb['Recognitions']

    # Find the first empty row
    row = ws.max_row + 1

    # Assuming recognition_data is a dictionary with column names and values
    for col_num, (col_name, value) in enumerate(recognition_data.items(), start=1):
        ws.cell(row=row, column=col_num, value=value)

    wb.save(file_path)
    print("New recognition data appended.")
