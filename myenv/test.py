import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import pandas as pd
import openpyxl 
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the Excel file
file_path = 'sample_order_data.xlsx'
print("File path:", os.path.abspath(file_path))


def adjust_column_widths(ws):
    for col in ws.iter_cols():
        max_length = 0
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width


# Specify the worksheet name where the data resides
def process_file(file_path):
    try:    
        wb = load_workbook(filename=file_path)
        worksheet_names = wb.sheetnames

        for worksheet_name in worksheet_names:
        # Read data from the worksheet
            df = pd.read_excel(file_path, sheet_name=worksheet_name, header=2) #added a header to start reading 2 rows down
            ws = wb[worksheet_name]
            adjust_column_widths(ws)

            # df edits  
            if worksheet_name == 'order_line_items':

                # Read required sheets
                products_df = pd.read_excel(file_path, sheet_name='products', header=2)
                order_line_items_df = pd.read_excel(file_path, sheet_name='order_line_items', header=2)
                orders_df = pd.read_excel(file_path, sheet_name='orders', header=2)
                customers_df = pd.read_excel(file_path, sheet_name='customers', header=2)
                salespersons_df = pd.read_excel(file_path, sheet_name='salespersons', header=2)

                print("Orders DataFrame:")
                print(orders_df.head())    
                print("Customers DataFrame:")
                print(customers_df.head()) 

                # Get the active worksheet
                ws = wb[worksheet_name]

            # Merge necessary DataFrames
                merged_price_df = pd.merge(order_line_items_df, products_df[['product_id', 'product_price']], on='product_id', how='left')
                # Assign the "product_price" values to the "item_price" column
                merged_price_df['item_price'] = merged_price_df['product_price']
                # Calculate total units
                merged_price_df['total_units'] = merged_price_df['quantity_ordered'] - merged_price_df['quantity_canceled']
                # Calculate line totals
                merged_price_df['line_total'] = merged_price_df['product_price'] * merged_price_df['total_units']
                # Calculate total line total for each order
                order_total_df = merged_price_df.groupby('order_id')['line_total'].agg('sum').reset_index()
                order_total_df['total_units'] = order_total_df['line_total']
            
            # Reasign columns to DataFrames
            
                order_line_items_df['item_price'] = merged_price_df['item_price']
                order_line_items_df['line_total'] = merged_price_df['line_total']
                order_line_items_df['total_units'] = merged_price_df['total_units']
                orders_df['order_total'] = order_total_df['line_total']       

            #Write data to sheets
                ws = wb['order_line_items']

                order_line_items_item_price = order_line_items_df.columns.get_loc('item_price')
                for idx, value in enumerate(order_line_items_df['item_price'], start=4):
                    ws.cell(row=idx, column=order_line_items_item_price + 1).value = value

                order_line_items_item_price = order_line_items_df.columns.get_loc('line_total')
                for idx, value in enumerate(order_line_items_df['line_total'], start=4):
                    ws.cell(row=idx, column=order_line_items_item_price + 1).value = value

                order_line_items_item_price = order_line_items_df.columns.get_loc('total_units')
                for idx, value in enumerate(order_line_items_df['total_units'], start=4):
                    ws.cell(row=idx, column=order_line_items_item_price + 1).value = value

                ws = wb['orders']
                orders_total_df = orders_df.columns.get_loc('order_total')
                for idx, value in enumerate(orders_df['order_total'], start=4):
                    ws.cell(row=idx, column=orders_total_df + 1).value = value    

                # Merge necessary DataFrames
                orders_df = pd.merge(order_line_items_df, orders_df, on='order_id',)

                orders_df['order_date'] = pd.to_datetime(orders_df['order_date'])
                order_totals = orders_df.groupby('order_id')['order_total'].sum().reset_index()

                monthly_sales = orders_df.groupby(orders_df['order_date'].dt.to_period('M'))['line_total'].sum()

                monthly_sales = monthly_sales.reset_index()

                monthly_sales.rename(columns={'order_date': 'date', 'line_total': 'sales'}, inplace=True)

                monthly_sales['date'] = monthly_sales['date'].dt.strftime('%Y-%m-%d')

                # Create or switch to Monthly_sales Sheet
                sheet_name = 'monthly_sales'
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(title=sheet_name)
                    
                ws = wb['monthly_sales']
                for r_idx, row in enumerate(dataframe_to_rows(monthly_sales, index=False, header=True), start=3):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value,)  

                for col in ws.iter_cols():
                    max_length = 0
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[col[0].column_letter].width = adjusted_width   
                
                # Create the 'total_spent' column
                ws = wb['customers']
                
                if 'total_spent' not in customers_df.columns:
                    customers_df['total_spent'] = 0
                            # Find the index of the last column
                    last_column_index = ws.max_column
                    # Insert a new column after the last column
                    ws.insert_cols(idx=last_column_index + 1)
                    # Set the header of the new column
                    ws.cell(row=3, column=last_column_index + 1, value='total_spent')

                customer_total_df = pd.merge(orders_df, customers_df, on='customer_id', how='left')

                total_spent_df = orders_df.groupby('customer_id')['line_total'].sum().reset_index()

                        # Rename columns for clarity
                total_spent_df.rename(columns={'line_total': 'total_spent'}, inplace=True)

                print("Orders DataFrame:")
                print(orders_df.head())  

                print("Total Spent DataFrame:")
                print(total_spent_df.head())    

                total_spent_column_index = customers_df.columns.get_loc('total_spent')
                for idx, value in enumerate(total_spent_df['total_spent'], start=4):
                    ws.cell(row=idx, column=total_spent_column_index + 1).value = value

        # Save the changes to the Excel file
        wb.save(file_path)
        return "Processing complete!"
    except Exception as e:
        return f"An error occurred: {e}"

def append_data(file_path, sheet_name, column_name, data):
    try:
        wb = load_workbook(filename=file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Read column names from the third row (index 2)
            column_names = [cell.value for cell in ws[3]]

            if column_name not in column_names:
                return f"Column '{column_name}' not found in sheet '{sheet_name}'."
            
            column_index = column_names.index(column_name) + 1

            # Find the first empty cell in the column starting from the 4th row
            for row in range(4, ws.max_row + 1):
                if ws.cell(row=row, column=column_index).value is None:
                    ws.cell(row=row, column=column_index).value = data
                    break
            else:
                # If no empty cell found, append to the end
                ws.cell(row=ws.max_row + 1, column=column_index).value = data

            wb.save(file_path)
            return "Data appended successfully!"
        else:
            return f"Sheet '{sheet_name}' not found."
    except Exception as e:
        return f"An error occurred: {e}"

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        file_path_label.config(text=os.path.basename(file_path))
        result = process_file(file_path)
        update_dropdowns(file_path)
        result_text.insert(tk.END, result + "\n")

def update_dropdowns(file_path):
    wb = load_workbook(filename=file_path)
    worksheet_names = wb.sheetnames
    sheet_name_combobox['values'] = worksheet_names
    sheet_name_combobox.current(0)

    update_column_names(file_path, worksheet_names[0])
    
def update_column_names(file_path, sheet_name):
    wb = load_workbook(filename=file_path)
    ws = wb[sheet_name]
    column_names = [cell.value for cell in ws[3]]
    column_name_combobox['values'] = column_names
    column_name_combobox.current(0)

def append_data_to_file():
    file_path = file_path_label.cget("text")
    if file_path == "No file selected":
        messagebox.showwarning("Warning", "Please select a file first.")
        return

    sheet_name = sheet_name_entry.get()
    column_name = column_name_entry.get()
    data = data_entry.get()

    if not sheet_name or not column_name or not data:
        messagebox.showwarning("Warning", "Please fill in all fields.")
        return

    result = append_data(file_path, sheet_name, column_name, data)
    result_text.insert(tk.END, result + "\n")

def add_new_sheet_or_column():
    file_path = file_path_label.cget("text")
    if file_path == "No file selected":
        messagebox.showwarning("Warning", "Please select a file first.")
        return

    new_sheet_name = new_sheet_name_entry.get()
    new_column_name = new_column_name_entry.get()

    if not new_sheet_name and not new_column_name:
        messagebox.showwarning("Warning", "Please provide a sheet name or a column name.")
        return

    wb = load_workbook(filename=file_path)

    if new_sheet_name:
        if new_sheet_name not in wb.sheetnames:
            wb.create_sheet(title=new_sheet_name)
        update_dropdowns(file_path)

    if new_column_name:
        sheet_name = sheet_name_combobox.get()
        ws = wb[sheet_name]
        column_names = [cell.value for cell in ws[3]]
        if new_column_name not in column_names:
            last_column_index = ws.max_column + 1
            ws.cell(row=3, column=last_column_index, value=new_column_name)
        update_column_names(file_path, sheet_name)

    wb.save(file_path)
    result_text.insert(tk.END, "New sheet or column added successfully!\n")

# Create the main window
root = tk.Tk()
root.title("Sample Data Processor")

# Create a frame for the file selection
frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

# Add a label and button for file selection
file_path_label = tk.Label(frame, text="No file selected")
file_path_label.pack(anchor="w")

select_file_button = tk.Button(frame, text="Select File", command=select_file)
select_file_button.pack()

# Add entry fields for appending data
entry_frame = tk.Frame(root)
entry_frame.pack(padx=10, pady=10)

tk.Label(frame, text="Sheet Name:").pack()
sheet_name_combobox = ttk.Combobox(frame, state="readonly")
sheet_name_combobox.pack()

tk.Label(frame, text="Column Name:").pack()
column_name_combobox = ttk.Combobox(frame, state="readonly")
column_name_combobox.pack()

tk.Label(frame, text="Data to Append:").pack()
data_entry = tk.Entry(frame)
data_entry.pack()

append_data_button = tk.Button(frame, text="Append Data", command=append_data_to_file)
append_data_button.pack()

    #Add these on another optional page
#tk.Label(frame, text="New Sheet Name:").pack()
#new_sheet_name_entry = tk.Entry(frame)
#new_sheet_name_entry.pack()

#tk.Label(frame, text="New Column Name:").pack()
#new_column_name_entry = tk.Entry(frame)
#new_column_name_entry.pack()

#add_new_button = tk.Button(frame, text="Add Sheet/Column", command=add_new_sheet_or_column)
#add_new_button.pack()

# Add a text widget for showing results
result_text = tk.Text(root, height=10, width=80)
result_text.pack(padx=10, pady=10)

# Start the GUI event loop
root.mainloop()